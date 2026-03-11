from flask import Blueprint, jsonify, send_file, Response
from services.alert_service import AlertService
from utils.auth_guard import check_role

import io
import csv
import random
from datetime import datetime

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

import openpyxl
from openpyxl.styles import Font, PatternFill

reports_routes = Blueprint("reports_routes", __name__)


@reports_routes.route("/reports/summary", methods=["GET"])
@check_role(["admin", "hr"])
def get_reports_summary():
    try:
        stats = AlertService.get_dashboard_stats()

        if not stats:
            return jsonify({"error": "No stats available"}), 500

        return jsonify({
            "total_employees": stats.get("total_employees", 0),
            "department_breakdown": [
                {"department": k, "count": v}
                for k, v in stats.get("dept_counts", {}).items()
            ],
            "risk_summary": {
                "on_track": stats.get("on_track", 0),
                "at_risk": stats.get("at_risk", 0),
                "delayed": stats.get("delayed", 0)
            },
            "averages": {
                "completion": stats.get("avg_completion", 0),
                "time_to_onboard": "14 days"
            },
            "top_risks": stats.get("critical_employees", [])[:5],
            "weekly_trend": _generate_trend_data()
        })

    except Exception as e:
        print("Error generating reports:", e)
        return jsonify({"error": "Failed to generate report"}), 500


def _generate_trend_data():
    try:
        user_risks = AlertService.get_user_risks() or {}

        total_users = len(user_risks)
        total_risk = 0

        for data in user_risks.values():
            status = data.get("status")

            if status == "On Track":
                total_risk += 10
            elif status == "At Risk":
                total_risk += 50
            elif status == "Delayed":
                total_risk += 90
            else:
                total_risk += 10

        current_risk_score = int(total_risk / total_users) if total_users else 0

        days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        trend_data = []

        base_risks = sum(
            1 for data in user_risks.values()
            if data.get("status") in ["At Risk", "Delayed"]
        )

        for i in range(6, -1, -1):
            variation = random.randint(-2, 2)
            day_risks = max(0, base_risks + variation)

            if i == 0:
                day_risks = base_risks

            day_label = days[(datetime.now().weekday() - i) % 7]

            trend_data.append({
                "day": day_label,
                "risks": day_risks
            })

        return trend_data

    except Exception as e:
        print("Trend generation error:", e)
        return []


@reports_routes.route("/reports/weekly-risk-trend", methods=["GET"])
@check_role(["admin", "hr"])
def get_weekly_risk_trend():
    try:
        return jsonify(_generate_trend_data())
    except Exception as e:
        print("Error generating risk trend:", e)
        return jsonify({"error": "Failed to generate risk trend"}), 500


# ==========================================
# PDF DOWNLOAD
# ==========================================

@reports_routes.route("/reports/download/pdf", methods=["GET"])
@check_role(["admin", "hr"])
def download_pdf():
    try:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)

        elements = []
        styles = getSampleStyleSheet()
        
        title_style = styles["Heading1"]
        title_style.alignment = 1 # Center
        subtitle_style = styles["Heading2"]
        subtitle_style.textColor = colors.HexColor("#2C3E50")
        subtitle_style.spaceAfter = 10
        normal_style = styles["Normal"]

        title = Paragraph("Onboarding Risk Analysis Report", title_style)
        elements.append(title)
        elements.append(Spacer(1, 12))

        elements.append(
            Paragraph(
                f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                styles["Normal"]
            )
        )

        elements.append(Spacer(1, 20))

        stats = AlertService.get_dashboard_stats()
        user_risks = AlertService.get_user_risks() or {}

        # 1. Summary Metrics
        elements.append(Paragraph("1. Summary Metrics", subtitle_style))
        overview_data = [
            ["Metric", "Value"],
            ["Total Employees", str(stats.get("total_employees", 0))],
            ["Avg Completion %", f"{stats.get('avg_completion',0)}%"],
            ["Time to Onboard Target", "14 Days"]
        ]
        
        summary_table = Table(overview_data, colWidths=[200, 100])
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#34495E")),
            ("TEXTCOLOR", (0,0), (-1,0), colors.whitesmoke),
            ("ALIGN", (0,0), (-1,-1), "LEFT"),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("BOTTOMPADDING", (0,0), (-1,0), 8),
            ("BACKGROUND", (0,1), (-1,-1), colors.HexColor("#ECF0F1")),
            ("GRID", (0,0), (-1,-1), 1, colors.white)
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))
        
        # 2. Employee Status Breakdown
        elements.append(Paragraph("2. Employee Status Breakdown", subtitle_style))
        status_data = [
            ["Status", "Count"],
            ["On Track", str(stats.get("on_track", 0))],
            ["At Risk", str(stats.get("at_risk", 0))],
            ["Delayed", str(stats.get("delayed", 0))]
        ]
        
        status_table = Table(status_data, colWidths=[200, 100])
        status_table.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#34495E")),
            ("TEXTCOLOR", (0,0), (-1,0), colors.whitesmoke),
            ("ALIGN", (0,0), (-1,-1), "LEFT"),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("BOTTOMPADDING", (0,0), (-1,0), 8),
            ("BACKGROUND", (0,1), (-1,-1), colors.HexColor("#ECF0F1")),
            ("GRID", (0,0), (-1,-1), 1, colors.white)
        ]))
        elements.append(status_table)
        elements.append(Spacer(1, 25))
        
        # Group users by status
        delayed_users = []
        at_risk_users = []
        on_track_users = []
        
        for data in user_risks.values():
            status = data.get("status")
            user = data.get("user")
            reasons_text = "; ".join(data.get("reasons", []))
            reasons_para = Paragraph(reasons_text, normal_style) if reasons_text else "N/A"
            row = [getattr(user, "name", ""), getattr(user, "department", "N/A"), reasons_para]
            
            if status == "Delayed":
                delayed_users.append(row)
            elif status == "At Risk":
                at_risk_users.append(row)
            else:
                on_track_users.append([getattr(user, "name", ""), getattr(user, "department", "N/A"), "N/A"])
                
        # 3. High Risk Employees (Delayed)
        elements.append(Paragraph("3. High Risk Employees (Delayed)", subtitle_style))
        if delayed_users:
            delayed_data = [["Employee Name", "Department", "Risk Reasons"]] + delayed_users
            delayed_table = Table(delayed_data, colWidths=[150, 120, 250])
            delayed_table.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#E74C3C")),
                ("TEXTCOLOR", (0,0), (-1,0), colors.whitesmoke),
                ("ALIGN", (0,0), (-1,-1), "LEFT"),
                ("VALIGN", (0,0), (-1,-1), "TOP"),
                ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
                ("BOTTOMPADDING", (0,0), (-1,0), 8),
                ("GRID", (0,0), (-1,-1), 1, colors.HexColor("#BDC3C7"))
            ]))
            elements.append(delayed_table)
        else:
            elements.append(Paragraph("No employees in 'Delayed' status.", normal_style))
        elements.append(Spacer(1, 25))
        
        # 4. At Risk Employees
        elements.append(Paragraph("4. At Risk Employees", subtitle_style))
        if at_risk_users:
            at_risk_data = [["Employee Name", "Department", "Risk Reasons"]] + at_risk_users
            at_risk_table = Table(at_risk_data, colWidths=[150, 120, 250])
            at_risk_table.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#F39C12")),
                ("TEXTCOLOR", (0,0), (-1,0), colors.whitesmoke),
                ("ALIGN", (0,0), (-1,-1), "LEFT"),
                ("VALIGN", (0,0), (-1,-1), "TOP"),
                ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
                ("BOTTOMPADDING", (0,0), (-1,0), 8),
                ("GRID", (0,0), (-1,-1), 1, colors.HexColor("#BDC3C7"))
            ]))
            elements.append(at_risk_table)
        else:
            elements.append(Paragraph("No employees in 'At Risk' status.", normal_style))
        elements.append(Spacer(1, 25))
        
        # 5. On Track Employees
        elements.append(Paragraph("5. On Track Employees", subtitle_style))
        if on_track_users:
            on_track_data = [["Employee Name", "Department", "Status"]] + on_track_users
            on_track_table = Table(on_track_data, colWidths=[150, 120, 250])
            on_track_table.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#27AE60")),
                ("TEXTCOLOR", (0,0), (-1,0), colors.whitesmoke),
                ("ALIGN", (0,0), (-1,-1), "LEFT"),
                ("VALIGN", (0,0), (-1,-1), "TOP"),
                ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
                ("BOTTOMPADDING", (0,0), (-1,0), 8),
                ("GRID", (0,0), (-1,-1), 1, colors.HexColor("#BDC3C7"))
            ]))
            elements.append(on_track_table)
        else:
            elements.append(Paragraph("No employees in 'On Track' status.", normal_style))

        doc.build(elements)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name="onboardai-report.pdf",
            mimetype="application/pdf"
        )

    except Exception as e:
        print("PDF generation error:", e)
        return jsonify({"error": "PDF generation failed"}), 500


# ==========================================
# CSV DOWNLOAD
# ==========================================

@reports_routes.route("/reports/download/csv", methods=["GET"])
@check_role(["admin", "hr"])
def download_csv():
    try:
        user_risks = AlertService.get_user_risks() or {}

        output = io.StringIO()
        writer = csv.writer(output)

        writer.writerow([
            "employee_id",
            "name",
            "email",
            "department",
            "role",
            "risk_status",
            "risk_reasons"
        ])

        for data in user_risks.values():
            user = data.get("user")

            writer.writerow([
                getattr(user, "id", ""),
                getattr(user, "name", ""),
                getattr(user, "email", ""),
                getattr(user, "department", ""),
                getattr(user, "role", ""),
                data.get("status"),
                "; ".join(data.get("reasons", []))
            ])

        output.seek(0)

        return Response(
            output.getvalue(),
            mimetype="text/csv",
            headers={
                "Content-Disposition":
                "attachment;filename=onboardai-report.csv"
            }
        )

    except Exception as e:
        print("CSV generation error:", e)
        return jsonify({"error": "CSV generation failed"}), 500


# ==========================================
# EXCEL DOWNLOAD
# ==========================================

@reports_routes.route("/reports/download/excel", methods=["GET"])
@check_role(["admin", "hr"])
def download_excel():
    try:
        user_risks = AlertService.get_user_risks() or {}
        stats = AlertService.get_dashboard_stats()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Employees"

        headers = [
            "employee_id",
            "name",
            "email",
            "department",
            "role",
            "risk_status",
            "risk_reasons"
        ]

        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="CCCCCC",
                end_color="CCCCCC",
                fill_type="solid"
            )

        for data in user_risks.values():
            user = data.get("user")

            ws.append([
                getattr(user, "id", ""),
                getattr(user, "name", ""),
                getattr(user, "email", ""),
                getattr(user, "department", ""),
                getattr(user, "role", ""),
                data.get("status"),
                "; ".join(data.get("reasons", []))
            ])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name="onboardai-report.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        print("Excel generation error:", e)
        return jsonify({"error": "Excel generation failed"}), 500